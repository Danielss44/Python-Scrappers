from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl
from itertools import zip_longest

# Entrar no site
# https://www.olx.com.br/brasil?q="nome do produtro"
urld = input("Busca: ")
driver = webdriver.Chrome()
driver.get(f'https://www.olx.com.br/brasil?q={urld}')


# Pegando os elementos

total = driver.find_elements(
    By.XPATH, "//div[@class='sc-c70b81f6-0 cUgHyT']/section")

titulo = driver.find_elements(
    By.XPATH, "//h2[@class='olx-text olx-text--title-small olx-text--block olx-ad-card__title olx-ad-card__title--horizontal']")

link = driver.find_elements(
    By.XPATH, "//div[@class='sc-c70b81f6-0 cUgHyT']/section/a")

dia = driver.find_elements(
    By.XPATH, "//div[@class='olx-ad-card__location-date-container']/p[@class='olx-text olx-text--caption olx-text--block olx-text--regular olx-ad-card__date--horizontal']")

estado = driver.find_elements(
    By.XPATH, "//p[@class='olx-text olx-text--caption olx-text--block olx-text--regular']")

preco = driver.find_elements(
    By.XPATH, "//div[@class='olx-ad-card__details-price--horizontal']/h3")

antigo = driver.find_elements(
    By.XPATH, "//p[@class='olx-text olx-text--caption olx-text--block olx-text--regular olx-ad-card__old-price olx-ad-card__old-price--horizontal']")

# Abrindo o arquivo

plan = openpyxl.load_workbook('produtos.xlsx')
planilha = plan['Plan1']

# Formatando o arquivo excel

for row in planilha.iter_rows(min_row=2, max_row=planilha.max_row, max_col=planilha.max_column):
    for cell in row:
        cell.value = None

# Listar com [Titulo, Link, Imagens, Data, Estado, Cidade, Preco ant, Preco atual]

for idx, (tit, lin, dat, est, prc, ant) in enumerate(zip_longest(titulo, link, dia, estado, preco, antigo, fillvalue=None), start=2):
    link_pronto = lin.get_attribute('href') if lin else "Sem link"
    titulo_texto = tit.text if tit else "Sem título"
    data_texto = dat.text if dat else "Sem data"
    cep_texto = est.text if est else "Sem dados"
    preco_texto = prc.text if prc else "Sem Preço"
    antigo_texto = ant.text if ant else "Sem Dados"

    cep_texto = cep_texto.split()
    cidade_texto = " ".join(
        cep_texto[:-2]) if len(cep_texto) > 2 else "Sem cidade"
    estado_texto = " ".join(
        cep_texto[-1:]) if len(cep_texto) >= 2 else "Sem estado"

    # Escrever diretamente nas células

    planilha.cell(row=idx, column=1).value = idx - \
        1  # Índice começa em 1 (linha 2, 3, ...)
    planilha.cell(row=idx, column=2).value = titulo_texto
    planilha.cell(row=idx, column=3).value = link_pronto
    planilha.cell(row=idx, column=4).value = data_texto
    planilha.cell(row=idx, column=5).value = cidade_texto
    planilha.cell(row=idx, column=6).value = estado_texto
    planilha.cell(row=idx, column=7).value = preco_texto
    planilha.cell(row=idx, column=8).value = antigo_texto

plan.save('produtos.xlsx')

driver.quit
